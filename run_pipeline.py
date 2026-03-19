import argparse
import asyncio
import csv
import logging
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from telethon import TelegramClient

import get_director_phone
import get_ip_phone
import get_phone_summary

load_dotenv()

DIRECT_PHONE_RE = re.compile(r"^7\d{10}$")

PIPELINE_FIELDNAMES = [
    "source_row",
    "source_name",
    "source_inn",
    "entity_type",
    "phone_source",
    "phone_lookup_status",
    "phone_lookup_message",
    "found_person",
    "found_phone",
    "found_email",
    "found_person_inn",
    "summary_status",
    "summary_message",
    "summary_fio",
    "summary_birth_date",
    "summary_age",
    "summary_telegram",
    "summary_email",
    "summary_inn",
    "vk_text",
    "vk_urls",
    "instagram_text",
    "instagram_urls",
    "ok_text",
    "ok_urls",
    "pipeline_status",
    "pipeline_message",
]

IP_MARKERS_RE = re.compile(r"\bИП\b|индивидуальн\w+\s+предпринимател\w+", re.IGNORECASE)
HEADER_NAME_MARKERS = ("название", "контрагент", "наименование", "company", "name")
HEADER_INN_MARKERS = ("инн", "inn")


@dataclass
class InputRow:
    source_row: int
    source_name: str
    source_inn: str | None


def setup_logging() -> logging.Logger:
    level_name = os.getenv("LOG_LEVEL", "INFO").strip().upper()
    level = getattr(logging, level_name, logging.INFO)
    logging.basicConfig(
        level=level,
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    )
    logging.getLogger("telethon").setLevel(logging.WARNING)
    return logging.getLogger("pipeline")


def get_required_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise RuntimeError(f"Missing required .env value: {name}")
    return value


def load_runtime_config() -> tuple[int, str, str, str, bool, Path, int, int]:
    api_id = int(get_required_env("API_ID"))
    api_hash = get_required_env("API_HASH")
    session_name = get_required_env("SESSION_NAME")
    bot_username = get_required_env("BOT")
    headless_raw = os.getenv("PLAYWRIGHT_HEADLESS", "1").strip().lower()
    headless = headless_raw not in {"0", "false", "no"}
    debug_dir = Path(os.getenv("REPORT_DEBUG_DIR", "report_debug").strip() or "report_debug")
    step_delay_seconds = int(os.getenv("PIPELINE_STEP_DELAY_SECONDS", "3").strip() or "3")
    row_delay_seconds = int(os.getenv("PIPELINE_ROW_DELAY_SECONDS", "5").strip() or "5")
    return api_id, api_hash, session_name, bot_username, headless, debug_dir, step_delay_seconds, row_delay_seconds


def normalize_inn(value: str | None) -> str | None:
    if not value:
        return None
    match = re.search(r"\b(\d{12}|\d{10})\b", str(value))
    if not match:
        return None
    return match.group(1)


def normalize_direct_phone(value: str | None) -> str | None:
    if value is None:
        return None
    candidate = str(value).strip()
    if not DIRECT_PHONE_RE.fullmatch(candidate):
        return None
    return candidate


def looks_like_header(first: str, second: str) -> bool:
    first_norm = first.strip().casefold()
    second_norm = second.strip().casefold()
    has_name = any(marker in first_norm for marker in HEADER_NAME_MARKERS)
    has_inn = any(marker in second_norm for marker in HEADER_INN_MARKERS)
    return has_name or has_inn


def detect_start_index(rows: list[tuple[str, str]]) -> int:
    if not rows:
        return 0

    first_name, first_inn = rows[0]
    if looks_like_header(first_name, first_inn):
        return 1

    if normalize_inn(first_inn):
        return 0

    if len(rows) > 1 and normalize_inn(rows[1][1]):
        return 1

    return 0


def iter_rows_from_csv(path: Path) -> Iterable[InputRow]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)

    if not rows:
        return []

    probe_rows = [
        (
            str(row[0]).strip() if len(row) > 0 and row[0] is not None else "",
            str(row[1]).strip() if len(row) > 1 and row[1] is not None else "",
        )
        for row in rows[:2]
    ]
    start_index = detect_start_index(probe_rows)
    out: list[InputRow] = []
    for index, row in enumerate(rows[start_index:], start=start_index + 1):
        if not row:
            continue
        source_name = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        source_inn = normalize_inn(str(row[1]).strip()) if len(row) > 1 and row[1] is not None else None
        if not source_name and not source_inn:
            continue
        out.append(InputRow(source_row=index, source_name=source_name, source_inn=source_inn))
    return out


def iter_rows_from_xlsx(path: Path) -> Iterable[InputRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        values = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not values:
        return []

    probe_rows = [
        (
            str(row[0]).strip() if len(row) > 0 and row[0] is not None else "",
            str(row[1]).strip() if len(row) > 1 and row[1] is not None else "",
        )
        for row in values[:2]
    ]
    start_index = detect_start_index(probe_rows)

    out: list[InputRow] = []
    for index, row in enumerate(values[start_index:], start=start_index + 1):
        row = row or ()
        source_name = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        source_inn = normalize_inn(str(row[1]).strip()) if len(row) > 1 and row[1] is not None else None
        if not source_name and not source_inn:
            continue
        out.append(InputRow(source_row=index, source_name=source_name, source_inn=source_inn))
    return out


def load_input_rows(path: Path) -> list[InputRow]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return list(iter_rows_from_csv(path))
    if suffix in {".xlsx", ".xlsm"}:
        return list(iter_rows_from_xlsx(path))
    raise RuntimeError(f"Unsupported input file format: {path.suffix}")


def detect_entity_type(source_name: str) -> str:
    if normalize_direct_phone(source_name):
        return "phone"
    return "ip" if IP_MARKERS_RE.search(source_name or "") else "company"


def append_result_csv(path: Path, row: dict[str, str | None]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not path.exists()
    with path.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=PIPELINE_FIELDNAMES)
        if write_header:
            writer.writeheader()
        writer.writerow(row)


def append_result_xlsx(path: Path, row: dict[str, str | None]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        workbook = load_workbook(path)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "pipeline"
        worksheet.append(PIPELINE_FIELDNAMES)

    worksheet.append([row.get(field) for field in PIPELINE_FIELDNAMES])
    workbook.save(path)


def append_pipeline_result(csv_path: Path, xlsx_path: Path, row: dict[str, str | None]) -> None:
    append_result_csv(csv_path, row)
    append_result_xlsx(xlsx_path, row)


def build_input_error_row(item: InputRow, message: str) -> dict[str, str | None]:
    return {
        "source_row": str(item.source_row),
        "source_name": item.source_name,
        "source_inn": item.source_inn,
        "entity_type": detect_entity_type(item.source_name),
        "phone_source": None,
        "phone_lookup_status": "input_error",
        "phone_lookup_message": message,
        "found_person": None,
        "found_phone": None,
        "found_email": None,
        "found_person_inn": None,
        "summary_status": None,
        "summary_message": None,
        "summary_fio": None,
        "summary_birth_date": None,
        "summary_age": None,
        "summary_telegram": None,
        "summary_email": None,
        "summary_inn": None,
        "vk_text": None,
        "vk_urls": None,
        "instagram_text": None,
        "instagram_urls": None,
        "ok_text": None,
        "ok_urls": None,
        "pipeline_status": "input_error",
        "pipeline_message": message,
    }


def build_pipeline_row(
    item: InputRow,
    *,
    entity_type: str,
    phone_source: str,
    phone_state,
    summary_state=None,
) -> dict[str, str | None]:
    phone_person = getattr(phone_state, "person", None)
    phone_lookup_status = phone_state.result_status
    phone_lookup_message = phone_state.status_message or phone_state.error
    found_phone = getattr(phone_person, "phone", None)

    summary = getattr(summary_state, "summary", None) if summary_state else None
    summary_status = summary_state.result_status if summary_state else None
    summary_message = (summary_state.status_message or summary_state.error) if summary_state else None

    if found_phone and summary_state and summary_state.result_status == "found":
        pipeline_status = "found"
        pipeline_message = "Телефон найден и краткая сводка получена"
    elif found_phone and summary_state:
        pipeline_status = "summary_failed"
        pipeline_message = summary_message
    elif found_phone:
        pipeline_status = "phone_found"
        pipeline_message = phone_lookup_message
    else:
        pipeline_status = "phone_lookup_failed"
        pipeline_message = phone_lookup_message

    return {
        "source_row": str(item.source_row),
        "source_name": item.source_name,
        "source_inn": item.source_inn,
        "entity_type": entity_type,
        "phone_source": phone_source,
        "phone_lookup_status": phone_lookup_status,
        "phone_lookup_message": phone_lookup_message,
        "found_person": getattr(phone_person, "fio", None),
        "found_phone": found_phone,
        "found_email": getattr(phone_person, "email", None),
        "found_person_inn": getattr(phone_person, "inn", None),
        "summary_status": summary_status,
        "summary_message": summary_message,
        "summary_fio": summary.fio if summary else None,
        "summary_birth_date": summary.birth_date if summary else None,
        "summary_age": summary.age if summary else None,
        "summary_telegram": summary.telegram if summary else None,
        "summary_email": summary.email if summary else None,
        "summary_inn": summary.inn if summary else None,
        "vk_text": summary.vk_text if summary else None,
        "vk_urls": summary.vk_urls if summary else None,
        "instagram_text": summary.instagram_text if summary else None,
        "instagram_urls": summary.instagram_urls if summary else None,
        "ok_text": summary.ok_text if summary else None,
        "ok_urls": summary.ok_urls if summary else None,
        "pipeline_status": pipeline_status,
        "pipeline_message": pipeline_message,
    }


def build_direct_phone_summary_row(
    item: InputRow,
    *,
    direct_phone: str,
    summary_state,
) -> dict[str, str | None]:
    summary = getattr(summary_state, "summary", None)
    summary_status = summary_state.result_status
    summary_message = summary_state.status_message or summary_state.error

    if summary_state.result_status == "found":
        pipeline_status = "found"
        pipeline_message = summary_message or "Phone summary received"
    else:
        pipeline_status = "summary_failed"
        pipeline_message = summary_message

    return {
        "source_row": str(item.source_row),
        "source_name": item.source_name,
        "source_inn": None,
        "entity_type": "phone",
        "phone_source": "direct_phone_input",
        "phone_lookup_status": None,
        "phone_lookup_message": None,
        "found_person": None,
        "found_phone": direct_phone,
        "found_email": None,
        "found_person_inn": None,
        "summary_status": summary_status,
        "summary_message": summary_message,
        "summary_fio": summary.fio if summary else None,
        "summary_birth_date": summary.birth_date if summary else None,
        "summary_age": summary.age if summary else None,
        "summary_telegram": summary.telegram if summary else None,
        "summary_email": summary.email if summary else None,
        "summary_inn": summary.inn if summary else None,
        "vk_text": summary.vk_text if summary else None,
        "vk_urls": summary.vk_urls if summary else None,
        "instagram_text": summary.instagram_text if summary else None,
        "instagram_urls": summary.instagram_urls if summary else None,
        "ok_text": summary.ok_text if summary else None,
        "ok_urls": summary.ok_urls if summary else None,
        "pipeline_status": pipeline_status,
        "pipeline_message": pipeline_message,
    }


async def resolve_row(
    client: TelegramClient,
    bot_entity,
    item: InputRow,
    *,
    log: logging.Logger,
    headless: bool,
    debug_dir: Path,
    step_delay_seconds: int,
) -> dict[str, str | None]:
    direct_phone = normalize_direct_phone(item.source_name)
    if direct_phone:
        summary_state = await get_phone_summary.run_single_query(
            client,
            bot_entity,
            direct_phone,
            log=log,
            persist=False,
            echo=False,
        )
        if step_delay_seconds > 0:
            await asyncio.sleep(step_delay_seconds)
        return build_direct_phone_summary_row(
            item,
            direct_phone=direct_phone,
            summary_state=summary_state,
        )
    if not item.source_inn:
        return build_input_error_row(item, "Во втором столбце не удалось распознать ИНН")

    entity_type = detect_entity_type(item.source_name)
    if entity_type == "ip":
        phone_source = "ip_web_flow"
        phone_state = await get_ip_phone.run_single_query(
            client,
            bot_entity,
            item.source_inn,
            log=log,
            persist=False,
            echo=False,
            headless=headless,
            debug_dir=debug_dir,
        )
    else:
        phone_source = "company_flow"
        phone_state = await get_director_phone.run_single_query(
            client,
            bot_entity,
            item.source_inn,
            log=log,
            persist=False,
            echo=False,
        )

    found_phone = getattr(getattr(phone_state, "person", None), "phone", None)
    if not found_phone:
        if step_delay_seconds > 0:
            await asyncio.sleep(step_delay_seconds)
        return build_pipeline_row(
            item,
            entity_type=entity_type,
            phone_source=phone_source,
            phone_state=phone_state,
        )

    if step_delay_seconds > 0:
        await asyncio.sleep(step_delay_seconds)

    summary_state = await get_phone_summary.run_single_query(
        client,
        bot_entity,
        found_phone,
        log=log,
        persist=False,
        echo=False,
    )

    if step_delay_seconds > 0:
        await asyncio.sleep(step_delay_seconds)

    return build_pipeline_row(
        item,
        entity_type=entity_type,
        phone_source=phone_source,
        phone_state=phone_state,
        summary_state=summary_state,
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run combined INN -> phone -> phone summary pipeline.")
    parser.add_argument("input_path", nargs="?", help="Path to input CSV/XLSX with name in column 1 and INN in column 2")
    parser.add_argument("--output-csv", dest="output_csv", default=os.getenv("PIPELINE_RESULTS_CSV", "pipeline_results.csv"))
    parser.add_argument("--output-xlsx", dest="output_xlsx", default=os.getenv("PIPELINE_RESULTS_XLSX", "pipeline_results.xlsx"))
    return parser.parse_args()


async def ainput(prompt: str) -> str:
    return await asyncio.to_thread(input, prompt)


async def main() -> None:
    args = parse_args()
    log = setup_logging()
    api_id, api_hash, session_name, bot_username, headless, debug_dir, step_delay_seconds, row_delay_seconds = load_runtime_config()

    input_path_raw = args.input_path
    if not input_path_raw:
        input_path_raw = (await ainput("Input file path (.csv/.xlsx): ")).strip()
    if not input_path_raw:
        raise RuntimeError("Input file path is required")

    input_path = Path(input_path_raw).expanduser()
    if not input_path.exists():
        raise RuntimeError(f"Input file not found: {input_path}")

    output_csv = Path(args.output_csv)
    output_xlsx = Path(args.output_xlsx)
    rows = load_input_rows(input_path)
    if not rows:
        raise RuntimeError(f"No data rows found in {input_path}")

    client = TelegramClient(session_name, api_id, api_hash)
    try:
        await client.connect()
        if not await client.is_user_authorized():
            raise RuntimeError(
                "Telegram session is not authorized. Run `python qr_login.py` first to create "
                f"`{session_name}.session`."
            )

        bot_entity = await client.get_entity(bot_username)
        log.info("Connected to bot %s", bot_username)
        print(f"Loaded {len(rows)} rows from {input_path}")
        print(f"Results -> {output_csv} and {output_xlsx}\n")

        for index, item in enumerate(rows, start=1):
            entity_type = detect_entity_type(item.source_name)
            if entity_type == "phone":
                print(f"[{index}/{len(rows)}] row={item.source_row} type={entity_type} phone={item.source_name}")
            else:
                print(f"[{index}/{len(rows)}] row={item.source_row} type={entity_type} inn={item.source_inn or 'missing'}")
            if item.source_name:
                print(f"    name: {item.source_name}")

            row = await resolve_row(
                client,
                bot_entity,
                item,
                log=log,
                headless=headless,
                debug_dir=debug_dir,
                step_delay_seconds=step_delay_seconds,
            )
            append_pipeline_result(output_csv, output_xlsx, row)

            print(f"    phone_lookup_status: {row['phone_lookup_status']}")
            print(f"    found_phone: {row['found_phone'] or 'not found'}")
            print(f"    summary_status: {row['summary_status'] or 'not run'}")
            print(f"    pipeline_status: {row['pipeline_status']}")
            print()

            if index < len(rows) and row_delay_seconds > 0:
                await asyncio.sleep(row_delay_seconds)
    finally:
        if client.is_connected():
            await client.disconnect()


if __name__ == "__main__":
    asyncio.run(main())
