import asyncio
import csv
import logging
import os
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from telethon import TelegramClient, events

load_dotenv()

COMPANY_INN_RE = re.compile(r"\bИНН\s*:\s*(\d{12}|\d{10})\b")
DIRECTOR_RE = re.compile(
    r"(?:Директор|Генеральный директор|Руководитель)\s*:\s*([^\n(]+?)(?:\s*\(ИНН\s*(\d{10,12})\))?(?:\n|$)"
)
FIO_RE = re.compile(r"\bФИО\s*:\s*([^\n]+)")
PHONE_RE = re.compile(r"\bТелефон\s*:\s*([+0-9][0-9()\-\s]{8,})")
EMAIL_RE = re.compile(r"\bEmail\s*:\s*([^\s\n]+)", re.IGNORECASE)
PERSON_INN_RE = re.compile(r"\bИНН\s*:\s*(\d{12}|\d{10})\b")
NOT_FOUND_RE = re.compile(r"к сожалению,\s*по данному запросу ничего не найдено", re.IGNORECASE)

CLICK_DELAY_SECONDS = 3
CLICK_TIMEOUT_SECONDS = 12
QUERY_TIMEOUT_SECONDS = 180
MAX_DEPTH = 5
RESULT_FIELDNAMES = [
    "requested_inn",
    "result_status",
    "status_message",
    "source_company_inn",
    "source_company_name",
    "last_company_inn",
    "last_company_name",
    "director_name",
    "director_inn",
    "person_fio",
    "phone",
    "email",
    "person_inn",
]


@dataclass
class CompanyCard:
    company_inn: str | None = None
    company_name: str | None = None
    director_name: str | None = None
    director_inn: str | None = None


@dataclass
class PersonCard:
    fio: str | None = None
    phone: str | None = None
    email: str | None = None
    inn: str | None = None
    raw_text: str | None = None


@dataclass
class ButtonCandidate:
    row_index: int
    col_index: int
    text: str


@dataclass
class QueryState:
    requested_inn: str
    queue: asyncio.Queue[Any] = field(default_factory=asyncio.Queue)
    source_company: CompanyCard | None = None
    last_company: CompanyCard | None = None
    person: PersonCard | None = None
    result_status: str = "pending"
    status_message: str | None = None
    error: str | None = None
    seen_cards: set[str] = field(default_factory=set)


def set_failure(
    state: QueryState,
    *,
    status: str,
    message: str,
    error: str | None = None,
) -> None:
    if state.result_status in {"found", "not_found"}:
        return
    state.result_status = status
    state.status_message = message
    state.error = error or message


def setup_logging() -> logging.Logger:
    level_name = os.getenv("LOG_LEVEL", "INFO").strip().upper()
    level = getattr(logging, level_name, logging.INFO)
    logging.basicConfig(
        level=level,
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    )
    logging.getLogger("telethon").setLevel(logging.WARNING)
    return logging.getLogger("director_phone")


def get_required_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise RuntimeError(f"Missing required .env value: {name}")
    return value


def load_config() -> tuple[int, str, str, str, Path, Path]:
    api_id = int(get_required_env("API_ID"))
    api_hash = get_required_env("API_HASH")
    session_name = get_required_env("SESSION_NAME")
    bot_username = get_required_env("BOT")
    results_csv = Path(os.getenv("RESULTS_CSV", "results.csv").strip() or "results.csv")
    default_xlsx = results_csv.with_suffix(".xlsx")
    results_xlsx = Path(os.getenv("RESULTS_XLSX", str(default_xlsx)).strip() or str(default_xlsx))
    return api_id, api_hash, session_name, bot_username, results_csv, results_xlsx


async def ainput(prompt: str) -> str:
    return await asyncio.to_thread(input, prompt)


def get_message_text(message) -> str:
    return (getattr(message, "raw_text", None) or getattr(message, "text", None) or "").strip()


def print_buttons(message) -> None:
    buttons = flatten_buttons(message)
    if not buttons:
        return
    print("[buttons]")
    for idx, button in enumerate(buttons, start=1):
        print(f"{idx}. {button.text}")


def print_incoming(prefix: str, message) -> None:
    text = get_message_text(message) or "[empty message]"
    print(f"\n{prefix} {text}")
    print_buttons(message)
    print("> ", end="", flush=True)


def extract_company_name(text: str) -> str | None:
    for line in text.splitlines():
        candidate = line.strip()
        if not candidate:
            continue
        if candidate.lower().startswith("/inn"):
            continue
        if candidate.startswith(("ИНН:", "ОГРН:", "Дата регистрации:", "Статус:", "Директор:")):
            break
        if candidate.startswith(("Финансовые показатели", "Адрес:", "Сотрудников:")):
            continue
        if re.match(r"^\d{2}\.\d{2}\b", candidate):
            continue
        if candidate.startswith("👁"):
            continue
        return candidate
    return None


def parse_company_card(text: str) -> CompanyCard | None:
    inn_match = COMPANY_INN_RE.search(text)
    director_match = DIRECTOR_RE.search(text)
    if not inn_match and not director_match:
        return None
    return CompanyCard(
        company_inn=inn_match.group(1) if inn_match else None,
        company_name=extract_company_name(text),
        director_name=director_match.group(1).strip() if director_match else None,
        director_inn=director_match.group(2).strip() if director_match and director_match.group(2) else None,
    )


def normalize_phone(raw_phone: str | None) -> str | None:
    if not raw_phone:
        return None
    cleaned = re.sub(r"[^\d+]", "", raw_phone)
    return cleaned or None


def parse_person_card(text: str) -> PersonCard | None:
    fio_match = FIO_RE.search(text)
    phone_match = PHONE_RE.search(text)
    email_match = EMAIL_RE.search(text)
    inn_match = PERSON_INN_RE.search(text)
    if not any((fio_match, phone_match, email_match)):
        return None
    return PersonCard(
        fio=fio_match.group(1).strip() if fio_match else None,
        phone=normalize_phone(phone_match.group(1)) if phone_match else None,
        email=email_match.group(1).strip() if email_match else None,
        inn=inn_match.group(1) if inn_match else None,
        raw_text=text,
    )


def parse_not_found_message(text: str) -> str | None:
    if NOT_FOUND_RE.search(text):
        return text.strip()
    return None


def normalize_inn(user_input: str) -> str | None:
    match = re.search(r"\b(\d{12}|\d{10})\b", user_input)
    if not match:
        return None
    return match.group(1)


def flatten_buttons(message) -> list[ButtonCandidate]:
    out: list[ButtonCandidate] = []
    rows = getattr(message, "buttons", None) or []
    for row_index, row in enumerate(rows):
        for col_index, button in enumerate(row):
            out.append(
                ButtonCandidate(
                    row_index=row_index,
                    col_index=col_index,
                    text=(button.text or "").strip(),
                )
            )
    return out


def should_skip_button(button_text: str) -> bool:
    normalized = " ".join(button_text.split()).casefold()
    return normalized.startswith("комментарии")


def build_result_row(state: QueryState) -> dict[str, str | None]:
    source_company = state.source_company or CompanyCard(company_inn=state.requested_inn)
    last_company = state.last_company

    return {
        "requested_inn": state.requested_inn,
        "result_status": state.result_status,
        "status_message": state.status_message,
        "source_company_inn": source_company.company_inn,
        "source_company_name": source_company.company_name,
        "last_company_inn": last_company.company_inn if last_company else None,
        "last_company_name": last_company.company_name if last_company else None,
        "director_name": last_company.director_name if last_company else None,
        "director_inn": last_company.director_inn if last_company else None,
        "person_fio": state.person.fio if state.person else None,
        "phone": state.person.phone if state.person else None,
        "email": state.person.email if state.person else None,
        "person_inn": state.person.inn if state.person else None,
    }


def append_result_csv(results_csv: Path, row: dict[str, str | None]) -> None:
    results_csv.parent.mkdir(parents=True, exist_ok=True)
    write_header = not results_csv.exists()
    with results_csv.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=RESULT_FIELDNAMES)
        if write_header:
            writer.writeheader()
        writer.writerow(row)


def append_result_xlsx(results_xlsx: Path, row: dict[str, str | None]) -> None:
    results_xlsx.parent.mkdir(parents=True, exist_ok=True)

    if results_xlsx.exists():
        workbook = load_workbook(results_xlsx)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "results"
        worksheet.append(RESULT_FIELDNAMES)

    worksheet.append([row.get(field) for field in RESULT_FIELDNAMES])
    workbook.save(results_xlsx)


def append_result(results_csv: Path, results_xlsx: Path, state: QueryState) -> None:
    row = build_result_row(state)
    append_result_csv(results_csv, row)
    append_result_xlsx(results_xlsx, row)


def drain_queue(state: QueryState) -> int:
    dropped = 0
    while True:
        try:
            state.queue.get_nowait()
            dropped += 1
        except asyncio.QueueEmpty:
            return dropped


async def wait_for_next_useful_message(
    state: QueryState,
    log: logging.Logger,
    timeout_seconds: int,
) -> tuple[str | None, Any | None, CompanyCard | PersonCard | None]:
    loop = asyncio.get_running_loop()
    deadline = loop.time() + timeout_seconds

    while True:
        remaining = deadline - loop.time()
        if remaining <= 0:
            return None, None, None

        try:
            message = await asyncio.wait_for(state.queue.get(), timeout=remaining)
        except asyncio.TimeoutError:
            return None, None, None

        text = get_message_text(message)
        not_found_message = parse_not_found_message(text)
        if not_found_message:
            log.info("Received not-found response for current query")
            return "not_found", message, not_found_message

        person = parse_person_card(text)
        if person:
            log.info("Received person card: fio=%s phone=%s", person.fio, person.phone)
            return "person", message, person

        company = parse_company_card(text)
        if company:
            log.info("Received company card: inn=%s name=%s", company.company_inn, company.company_name)
            return "company", message, company

        log.debug("Skip bot message: not a company/person card")


async def explore_message(
    message,
    state: QueryState,
    log: logging.Logger,
    *,
    depth: int,
) -> bool:
    if depth > MAX_DEPTH:
        set_failure(
            state,
            status="max_depth_exceeded",
            message=f"Превышена максимальная глубина обхода ({MAX_DEPTH})",
        )
        return False

    text = get_message_text(message)
    card_key = text.strip()
    if card_key in state.seen_cards:
        print(f"{'  ' * depth}[trace] repeated card skipped")
        return False
    state.seen_cards.add(card_key)

    person = parse_person_card(text)
    if person:
        if person.phone:
            state.person = person
            print(f"{'  ' * depth}[trace] person found: {person.fio or 'unknown'}")
            return True
        print(f"{'  ' * depth}[warn] person card without phone")
        return False

    company = parse_company_card(text)
    if not company:
        print(f"{'  ' * depth}[warn] unsupported card format")
        return False

    if state.source_company is None:
        state.source_company = company
    state.last_company = company

    indent = "  " * depth
    company_label = company.company_name or company.company_inn or "unknown company"
    print(f"{indent}[trace] company card: {company_label}")

    buttons = flatten_buttons(message)
    if not buttons:
        print(f"{indent}[warn] no buttons on this card")
        return False

    print(f"{indent}[buttons] found {len(buttons)}:")
    for idx, button in enumerate(buttons, start=1):
        print(f"{indent}{idx}. {button.text}")

    actionable_buttons = [button for button in buttons if not should_skip_button(button.text)]
    skipped_buttons = [button for button in buttons if should_skip_button(button.text)]

    for button in skipped_buttons:
        print(f"{indent}[skip] service button: {button.text}")

    if not actionable_buttons:
        print(f"{indent}[warn] no actionable buttons on this card")
        return False

    for idx, button in enumerate(actionable_buttons, start=1):
        dropped = drain_queue(state)
        if dropped:
            log.debug("Dropped %s stale queued messages before click", dropped)

        print(
            f"{indent}[click] wait {CLICK_DELAY_SECONDS}s before button "
            f"{idx}/{len(actionable_buttons)}: {button.text}"
        )
        await asyncio.sleep(CLICK_DELAY_SECONDS)

        try:
            await message.click(button.row_index, button.col_index)
        except Exception as exc:
            print(f"{indent}[warn] click failed: {button.text} ({exc})")
            log.exception("Click failed for button %s", button.text)
            continue

        print(f"{indent}[click] pressed: {button.text}")

        kind, next_message, payload = await wait_for_next_useful_message(
            state,
            log,
            timeout_seconds=CLICK_TIMEOUT_SECONDS,
        )
        if kind is None:
            print(f"{indent}[warn] no useful response for button: {button.text}")
            continue

        if kind == "person":
            person = payload
            assert isinstance(person, PersonCard)
            if person.phone:
                state.person = person
                print(f"{indent}[trace] phone found: {person.phone}")
                return True
            print(f"{indent}[warn] person card without phone after button: {button.text}")
            continue

        next_company = payload
        assert isinstance(next_company, CompanyCard)
        print(
            f"{indent}[trace] nested company after button '{button.text}': "
            f"{next_company.company_name or next_company.company_inn or 'unknown company'}"
        )
        if await explore_message(next_message, state, log, depth=depth + 1):
            return True

        print(f"{indent}[trace] branch ended without phone: {button.text}")

    return False


async def resolve_query(state: QueryState, log: logging.Logger) -> bool:
    kind, message, payload = await wait_for_next_useful_message(
        state,
        log,
        timeout_seconds=CLICK_TIMEOUT_SECONDS,
    )
    if kind is None:
        set_failure(
            state,
            status="no_response",
            message="После /inn бот не вернул полезный ответ",
        )
        return False

    if kind == "not_found":
        not_found_message = payload
        assert isinstance(not_found_message, str)
        set_failure(
            state,
            status="not_found",
            message=not_found_message,
            error="По этому ИНН бот ничего не нашёл",
        )
        return False

    if kind == "person":
        person = payload
        assert isinstance(person, PersonCard)
        if not person.phone:
            set_failure(
                state,
                status="phone_not_found",
                message="Карточка физлица получена, но телефон отсутствует",
            )
            return False
        state.person = person
        state.result_status = "found"
        return True

    assert message is not None
    found = await explore_message(message, state, log, depth=0)
    if found:
        state.result_status = "found"
    elif state.result_status == "pending":
        set_failure(
            state,
            status="phone_not_found",
            message="Телефон не найден после обхода всех кнопок",
        )
    return found


async def run_single_query(
    client: TelegramClient,
    bot_entity,
    inn: str,
    *,
    log: logging.Logger | None = None,
    results_csv: Path | None = None,
    results_xlsx: Path | None = None,
    persist: bool = False,
    echo: bool = True,
    timeout_seconds: int = QUERY_TIMEOUT_SECONDS,
) -> QueryState:
    query_log = log or logging.getLogger("director_phone")
    state = QueryState(requested_inn=inn)

    async def handle_bot_message(event, prefix: str) -> None:
        message = event.message
        if echo:
            print_incoming(prefix, message)
        state.queue.put_nowait(message)

    async def on_new_message(event):
        await handle_bot_message(event, "<")

    async def on_edited_message(event):
        await handle_bot_message(event, "< [edit]")

    new_message_builder = events.NewMessage(from_users=bot_entity)
    edited_message_builder = events.MessageEdited(from_users=bot_entity)
    client.add_event_handler(on_new_message, new_message_builder)
    client.add_event_handler(on_edited_message, edited_message_builder)

    try:
        command = f"/inn {inn}"
        if echo:
            print(f"[you] {command}")
        await client.send_message(bot_entity, command)

        try:
            found = await asyncio.wait_for(resolve_query(state, query_log), timeout=timeout_seconds)
        except asyncio.TimeoutError:
            set_failure(
                state,
                status="timeout",
                message=f"Превышено время ожидания результата ({timeout_seconds} сек)",
            )
        else:
            if found and state.result_status == "pending":
                state.result_status = "found"

        if persist:
            if results_csv is None or results_xlsx is None:
                raise RuntimeError("results_csv/results_xlsx are required when persist=True")
            append_result(results_csv, results_xlsx, state)

        return state
    finally:
        client.remove_event_handler(on_new_message, new_message_builder)
        client.remove_event_handler(on_edited_message, edited_message_builder)


async def main() -> None:
    log = setup_logging()
    api_id, api_hash, session_name, bot_username, results_csv, results_xlsx = load_config()
    client = TelegramClient(session_name, api_id, api_hash)
    current_query: QueryState | None = None

    try:
        await client.connect()
        if not await client.is_user_authorized():
            raise RuntimeError(
                "Telegram session is not authorized. Run `python qr_login.py` first to create "
                f"`{session_name}.session`."
            )

        bot_entity = await client.get_entity(bot_username)
        log.info("Connected to bot %s", bot_username)

        async def handle_bot_message(event, prefix: str) -> None:
            nonlocal current_query

            message = event.message
            print_incoming(prefix, message)

            if current_query is not None:
                current_query.queue.put_nowait(message)

        @client.on(events.NewMessage(from_users=bot_entity))
        async def on_new_message(event):
            await handle_bot_message(event, "<")

        @client.on(events.MessageEdited(from_users=bot_entity))
        async def on_edited_message(event):
            await handle_bot_message(event, "< [edit]")

        print("Ready. Enter INN or /exit.\n")

        while True:
            user_input = (await ainput("> ")).strip()
            if not user_input:
                continue
            if user_input == "/exit":
                break

            inn = normalize_inn(user_input)
            if not inn:
                print("[warn] Could not parse INN from input")
                continue

            current_query = QueryState(requested_inn=inn)
            command = f"/inn {inn}"
            print(f"[you] {command}")
            await client.send_message(bot_entity, command)

            try:
                found = await asyncio.wait_for(resolve_query(current_query, log), timeout=QUERY_TIMEOUT_SECONDS)
            except asyncio.TimeoutError:
                set_failure(
                    current_query,
                    status="timeout",
                    message=f"Превышено время ожидания результата ({QUERY_TIMEOUT_SECONDS} сек)",
                )
                append_result(results_csv, results_xlsx, current_query)
                print("\n[result]")
                print(f"requested_inn: {current_query.requested_inn}")
                print("status: timeout")
                print(f"message: {current_query.status_message}")
                print(f"saved_to: {results_csv}")
                print(f"saved_to: {results_xlsx}")
                print()
                current_query = None
                continue

            if not found:
                append_result(results_csv, results_xlsx, current_query)
                print("\n[result]")
                print(f"requested_inn: {current_query.requested_inn}")
                print(f"status: {current_query.result_status}")
                print(f"message: {current_query.status_message or current_query.error or 'Unknown error'}")
                print(f"saved_to: {results_csv}")
                print(f"saved_to: {results_xlsx}")
                print()
                current_query = None
                continue

            append_result(results_csv, results_xlsx, current_query)

            print("\n[result]")
            print(f"requested_inn: {current_query.requested_inn}")
            print(
                f"source_company: "
                f"{(current_query.source_company.company_name if current_query.source_company else '') or 'unknown'}"
            )
            print(f"person: {current_query.person.fio or 'unknown'}")
            print(f"phone: {current_query.person.phone or 'not found'}")
            print(f"email: {current_query.person.email or 'not found'}")
            print(f"saved_to: {results_csv}")
            print(f"saved_to: {results_xlsx}")
            print()

            current_query = None
    finally:
        if client.is_connected():
            await client.disconnect()


if __name__ == "__main__":
    asyncio.run(main())
