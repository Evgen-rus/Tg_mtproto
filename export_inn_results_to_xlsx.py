import json
import os
import sqlite3
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def _default_out_path() -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"inn_results_{ts}.xlsx"


def _connect(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def _founders_pretty(founders_json: str | None) -> str | None:
    if not founders_json:
        return None
    try:
        data = json.loads(founders_json)
    except json.JSONDecodeError:
        return founders_json

    if not isinstance(data, list):
        return founders_json

    parts: list[str] = []
    for item in data:
        if not isinstance(item, dict):
            parts.append(str(item))
            continue
        name = item.get("name") or ""
        inn = item.get("inn") or ""
        share = item.get("share_percent")
        if share is None:
            parts.append(f"{name} (ИНН {inn})".strip())
        else:
            parts.append(f"{name} (ИНН {inn}, {share}%)".strip())
    return "\n".join([p for p in parts if p.strip()]) or None


def fetch_rows(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute(
        """
        SELECT
            r.inn,
            r.ogrn,
            r.company_name,
            r.okved,
            r.reg_date,
            r.company_status,
            r.director_name,
            r.director_inn,
            r.employees_count,
            r.revenue_2024,
            r.income_2024,
            r.expenses_2024,
            r.authorized_capital,
            r.address,
            r.founders_json,
            r.source_query_id,
            q.query_text AS source_query_text,
            q.created_at AS source_query_created_at,
            r.created_at,
            r.updated_at,
            r.raw_text
        FROM inn_results r
        JOIN source_queries q ON q.id = r.source_query_id
        ORDER BY r.updated_at DESC
        ;
        """
    ).fetchall()


def autosize_columns(ws) -> None:
    # Простое авторасширение (без дорогостоящих измерений шрифтов).
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 30)


def export_to_xlsx(
    rows: list[sqlite3.Row],
    out_path: str,
    *,
    include_founders_pretty: bool = True,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "inn_results"

    headers = [
        "ИНН",
        "ОГРН",
        "Название",
        "ОКВЭД",
        "Дата регистрации",
        "Статус",
        "Директор (ФИО)",
        "Директор (ИНН)",
        "Сотрудников",
        "Выручка 2024",
        "Доход 2024",
        "Расходы 2024",
        "Уставный капитал",
        "Адрес",
        "Учредители",
        "Исходный запрос",
        "Создано (UTC)",
        "raw_text",
    ]

    ws.append(headers)

    for row in rows:
        founders_json = row["founders_json"]
        values = [
            row["inn"],
            row["ogrn"],
            row["company_name"],
            row["okved"],
            row["reg_date"],
            row["company_status"],
            row["director_name"],
            row["director_inn"],
            row["employees_count"],
            row["revenue_2024"],
            row["income_2024"],
            row["expenses_2024"],
            row["authorized_capital"],
            row["address"],
            _founders_pretty(founders_json),
            row["source_query_text"],
            row["created_at"],
            row["raw_text"],
        ]
        ws.append(values)

    # Оформление
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    autosize_columns(ws)

    # Ограничить высоту строк (особенно для raw_text) до ~5 строк текста (~100pt)
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 100

    wb.save(out_path)


def main() -> int:
    # DB_PATH можно задать в .env (как и в 2_mtproto.py).
    db_path = os.getenv("DB_PATH", "tg_results.db").strip()
    out_path = _default_out_path()

    if not os.path.exists(db_path):
        print(f"[error] DB file not found: {db_path}")
        return 2

    conn = _connect(db_path)
    try:
        rows = fetch_rows(conn)
    finally:
        conn.close()

    export_to_xlsx(
        rows,
        out_path,
        include_founders_pretty=True,
    )
    print(f"[ok] Exported {len(rows)} rows to: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

