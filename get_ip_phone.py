import asyncio
import csv
import logging
import os
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from playwright.async_api import TimeoutError as PlaywrightTimeoutError
from playwright.async_api import async_playwright
from telethon import TelegramClient, events

from telethon_client_factory import build_telegram_client

load_dotenv()

NOT_FOUND_RE = re.compile(r"к сожалению,\s*по данному запросу ничего не найдено", re.IGNORECASE)
REPORT_LINK_RE = re.compile(r"полным отчетом можно по ссылке ниже", re.IGNORECASE)

PHONE_RE = re.compile(r"(?:Телефон|РўРµР»РµС„РѕРЅ)(?:\s*:)?\s*([+0-9][0-9()\-\s]{8,})", re.IGNORECASE)
EMAIL_RE = re.compile(r"Email(?:\s*:)?\s*([^\s\n]+)", re.IGNORECASE)
PERSON_INN_RE = re.compile(r"\b(?:ИНН|РРќРќ)(?:\s*:)?\s*(\d{12}|\d{10})\b")

SUMMARY_MARKERS = ("Краткая сводка", "РљСЂР°С‚РєР°СЏ СЃРІРѕРґРєР°")
PERSONS_MARKERS = ("Личности", "Р›РёС‡РЅРѕСЃС‚Рё")
SUMMARY_END_MARKERS = (
    "Отчёты по найденным лицам",
    "Отчеты по найденным лицам",
    "Расширенный отчет",
    "Расширенный отчёт",
    "Место работы",
    "Связанные компании",
    "Адреса",
    "Интернет активность",
    "РћС‚С‡С‘С‚С‹ РїРѕ РЅР°Р№РґРµРЅРЅС‹Рј Р»РёС†Р°Рј",
    "РћС‚С‡РµС‚С‹ РїРѕ РЅР°Р№РґРµРЅРЅС‹Рј Р»РёС†Р°Рј",
    "Р Р°СЃС€РёСЂРµРЅРЅС‹Р№ РѕС‚С‡РµС‚",
    "Р Р°СЃС€РёСЂРµРЅРЅС‹Р№ РѕС‚С‡С‘С‚",
    "РњРµСЃС‚Рѕ СЂР°Р±РѕС‚С‹",
    "РЎРІСЏР·Р°РЅРЅС‹Рµ РєРѕРјРїР°РЅРёРё",
    "РђРґСЂРµСЃР°",
    "РРЅС‚РµСЂРЅРµС‚ Р°РєС‚РёРІРЅРѕСЃС‚СЊ",
)

RESULT_FIELDNAMES = [
    "requested_inn",
    "result_status",
    "status_message",
    "source_company_inn",
    "source_company_name",
    "last_company_inn",
    "last_company_name",
    "director_name",
    "director_inn",
    "person_fio",
    "phone",
    "email",
    "person_inn",
]

QUERY_TIMEOUT_SECONDS = 90
PAGE_TIMEOUT_SECONDS = 60000


@dataclass
class ReportLinkCandidate:
    row_index: int
    col_index: int
    text: str
    url: str | None


@dataclass
class WebPersonResult:
    fio: str | None = None
    phone: str | None = None
    email: str | None = None
    inn: str | None = None
    report_url: str | None = None


@dataclass
class QueryState:
    requested_inn: str
    queue: asyncio.Queue[Any] = field(default_factory=asyncio.Queue)
    result_status: str = "pending"
    status_message: str | None = None
    person: WebPersonResult | None = None
    error: str | None = None


def setup_logging() -> logging.Logger:
    level_name = os.getenv("LOG_LEVEL", "INFO").strip().upper()
    level = getattr(logging, level_name, logging.INFO)
    logging.basicConfig(
        level=level,
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    )
    logging.getLogger("telethon").setLevel(logging.WARNING)
    return logging.getLogger("ip_phone")


def get_required_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise RuntimeError(f"Missing required .env value: {name}")
    return value


def get_bool_env(name: str, default: bool) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    return raw.strip().lower() not in {"0", "false", "no", "off"}


def load_config() -> tuple[int, str, str, str, Path, Path, bool, Path, bool]:
    api_id = int(get_required_env("API_ID"))
    api_hash = get_required_env("API_HASH")
    session_name = get_required_env("SESSION_NAME")
    bot_username = get_required_env("BOT")
    results_csv = Path(os.getenv("RESULTS_CSV", "results.csv").strip() or "results.csv")
    default_xlsx = results_csv.with_suffix(".xlsx")
    results_xlsx = Path(os.getenv("RESULTS_XLSX", str(default_xlsx)).strip() or str(default_xlsx))
    headless = get_bool_env("PLAYWRIGHT_HEADLESS", True)
    debug_dir = Path(os.getenv("REPORT_DEBUG_DIR", "report_debug").strip() or "report_debug")
    bot_message_echo = get_bool_env("BOT_MESSAGE_ECHO", True)
    return api_id, api_hash, session_name, bot_username, results_csv, results_xlsx, headless, debug_dir, bot_message_echo


def set_failure(state: QueryState, *, status: str, message: str, error: str | None = None) -> None:
    if state.result_status == "found":
        return
    state.result_status = status
    state.status_message = message
    state.error = error or message


async def ainput(prompt: str) -> str:
    return await asyncio.to_thread(input, prompt)


def normalize_inn(user_input: str) -> str | None:
    match = re.search(r"\b(\d{12}|\d{10})\b", user_input)
    if not match:
        return None
    return match.group(1)


def get_message_text(message) -> str:
    return (getattr(message, "raw_text", None) or getattr(message, "text", None) or "").strip()


def should_skip_button(button_text: str) -> bool:
    normalized = " ".join(button_text.split()).casefold()
    return normalized.startswith("комментарии")


def flatten_buttons(message) -> list[ReportLinkCandidate]:
    out: list[ReportLinkCandidate] = []
    rows = getattr(message, "buttons", None) or []
    for row_index, row in enumerate(rows):
        for col_index, button in enumerate(row):
            out.append(
                ReportLinkCandidate(
                    row_index=row_index,
                    col_index=col_index,
                    text=(button.text or "").strip(),
                    url=getattr(button, "url", None),
                )
            )
    return out


def print_buttons(message) -> None:
    buttons = flatten_buttons(message)
    if not buttons:
        return
    print("[buttons]")
    for idx, button in enumerate(buttons, start=1):
        suffix = f" | url={button.url}" if button.url else ""
        print(f"{idx}. row={button.row_index} col={button.col_index} text={button.text}{suffix}")


def print_incoming(prefix: str, message) -> None:
    text = get_message_text(message) or "[empty message]"
    print(f"\n{prefix} {text}")
    print_buttons(message)
    print("> ", end="", flush=True)


def parse_not_found_message(text: str) -> str | None:
    if NOT_FOUND_RE.search(text):
        return text.strip()
    return None


def extract_report_button(message) -> ReportLinkCandidate | None:
    buttons = flatten_buttons(message)
    actionable_buttons = [button for button in buttons if not should_skip_button(button.text)]
    if not actionable_buttons:
        return None

    print(f"[buttons] found {len(buttons)}:")
    for idx, button in enumerate(buttons, start=1):
        suffix = f" | url={button.url}" if button.url else ""
        print(f"{idx}. row={button.row_index} col={button.col_index} text={button.text}{suffix}")

    for button in buttons:
        if should_skip_button(button.text):
            print(f"[skip] service button: {button.text}")

    for button in actionable_buttons:
        if button.url:
            print(f"[select] report button: {button.text}")
            return button

    return None


async def wait_for_report_message(state: QueryState, log: logging.Logger) -> tuple[str | None, Any | None]:
    loop = asyncio.get_running_loop()
    deadline = loop.time() + QUERY_TIMEOUT_SECONDS

    while True:
        remaining = deadline - loop.time()
        if remaining <= 0:
            return None, None

        try:
            message = await asyncio.wait_for(state.queue.get(), timeout=remaining)
        except asyncio.TimeoutError:
            return None, None

        text = get_message_text(message)
        not_found_message = parse_not_found_message(text)
        if not_found_message:
            log.info("Received not-found response for current query")
            return "not_found", not_found_message

        if REPORT_LINK_RE.search(text):
            log.info("Received report-link response for current query")
            return "report_link", message


def normalize_phone(raw_phone: str | None) -> str | None:
    if not raw_phone:
        return None
    cleaned = re.sub(r"[^\d+]", "", raw_phone)
    if cleaned.startswith("+7") and len(cleaned) > 12:
        cleaned = cleaned[:12]
    elif cleaned.startswith("7") and len(cleaned) > 11:
        cleaned = cleaned[:11]
    elif cleaned.startswith("8") and len(cleaned) > 11:
        cleaned = cleaned[:11]
    return cleaned or None


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def maybe_fix_mojibake(text: str) -> str:
    if not text:
        return text

    mojibake_markers = ("Р", "С", "вЂ", "рџ")
    if not any(marker in text for marker in mojibake_markers):
        return text

    try:
        fixed = text.encode("cp1251", errors="ignore").decode("utf-8", errors="ignore")
    except Exception:
        return text

    # Accept only if the repaired text looks more like readable Russian.
    original_score = sum(text.count(marker) for marker in mojibake_markers)
    fixed_score = sum(fixed.count(marker) for marker in mojibake_markers)
    if fixed and fixed_score < original_score:
        return fixed
    return text


def extract_summary_section(body_text: str) -> str:
    summary = body_text
    for marker in SUMMARY_MARKERS:
        if marker in body_text:
            summary = body_text.split(marker, 1)[1]
            break

    for marker in SUMMARY_END_MARKERS:
        if marker in summary:
            summary = summary.split(marker, 1)[0]
    return summary.strip()


def extract_first_person_line(summary_text: str) -> str | None:
    lines = [normalize_space(line) for line in summary_text.splitlines() if normalize_space(line)]
    for idx, line in enumerate(lines):
        if any(line.casefold() == marker.casefold() for marker in PERSONS_MARKERS):
            for candidate in lines[idx + 1 :]:
                if re.search(r"\d{2}\.\d{2}\.\d{4}", candidate):
                    return candidate
    return None


def clean_person_name(raw_line: str | None) -> str | None:
    if not raw_line:
        return None
    line = re.sub(r"\s+\d+(?:[.,]\d+)?%$", "", raw_line).strip()
    match = re.match(r"(.+?)(?:\s+\d{2}\.\d{2}\.\d{4})?$", line)
    if not match:
        return line
    return match.group(1).strip()


def parse_report_text_once(body_text: str, report_url: str) -> WebPersonResult | None:
    summary = extract_summary_section(body_text)
    search_text = summary if normalize_space(summary) else body_text
    first_person_line = extract_first_person_line(search_text)
    phone_match = PHONE_RE.search(search_text)
    email_match = EMAIL_RE.search(search_text)
    inn_match = PERSON_INN_RE.search(search_text)

    result = WebPersonResult(
        fio=clean_person_name(first_person_line),
        phone=normalize_phone(phone_match.group(1)) if phone_match else None,
        email=email_match.group(1).strip() if email_match else None,
        inn=inn_match.group(1) if inn_match else None,
        report_url=report_url,
    )

    if not any((result.fio, result.phone, result.email, result.inn)):
        return None
    return result


def parse_report_text(body_text: str, report_url: str) -> WebPersonResult | None:
    candidates: list[str] = [body_text]

    fixed_body_text = maybe_fix_mojibake(body_text)
    if fixed_body_text != body_text:
        candidates.append(fixed_body_text)

    for candidate in candidates:
        parsed = parse_report_text_once(candidate, report_url)
        if parsed and (parsed.phone or parsed.fio or parsed.inn):
            return parsed

    return None


def save_report_debug_artifacts(
    debug_dir: Path,
    *,
    requested_inn: str,
    page_title: str,
    page_url: str,
    body_text: str,
    fixed_body_text: str,
    html_text: str,
) -> tuple[Path, Path, Path | None]:
    debug_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = f"{requested_inn}_{ts}"
    body_path = debug_dir / f"{stem}_body.txt"
    html_path = debug_dir / f"{stem}_page.html"
    fixed_body_path: Path | None = None

    body_dump = (
        f"requested_inn: {requested_inn}\n"
        f"page_url: {page_url}\n"
        f"page_title: {page_title}\n"
        f"body_len: {len(body_text)}\n"
        "----- BODY TEXT -----\n"
        f"{body_text}"
    )
    body_path.write_text(body_dump, encoding="utf-8")
    html_path.write_text(html_text, encoding="utf-8")
    if fixed_body_text != body_text:
        fixed_body_path = debug_dir / f"{stem}_fixed_body.txt"
        fixed_dump = (
            f"requested_inn: {requested_inn}\n"
            f"page_url: {page_url}\n"
            f"page_title: {page_title}\n"
            f"body_len: {len(fixed_body_text)}\n"
            "----- FIXED BODY TEXT -----\n"
            f"{fixed_body_text}"
        )
        fixed_body_path.write_text(fixed_dump, encoding="utf-8")
    return body_path, html_path, fixed_body_path


async def fetch_person_from_report(
    url: str,
    *,
    requested_inn: str,
    headless: bool,
    debug_dir: Path,
    log: logging.Logger,
) -> tuple[WebPersonResult | None, Path | None, Path | None, Path | None]:
    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=headless)
        page = await browser.new_page()
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT_SECONDS)
            try:
                await page.wait_for_load_state("networkidle", timeout=10000)
            except PlaywrightTimeoutError:
                log.debug("networkidle not reached, continue with current page state")
            try:
                await page.wait_for_selector("text=Краткая сводка", timeout=10000)
            except PlaywrightTimeoutError:
                log.debug("text 'Краткая сводка' not found, continue with current page state")
            await page.wait_for_timeout(2000)
            page_title = await page.title()
            page_url = page.url
            body_text = await page.locator("body").inner_text()
            html_text = await page.content()
            fixed_body_text = maybe_fix_mojibake(body_text)
            fixed_page_title = maybe_fix_mojibake(page_title)

            log.info(
                "Report page loaded: title=%r url=%s body_len=%s",
                fixed_page_title,
                page_url,
                len(fixed_body_text),
            )

            parsed = parse_report_text(body_text, report_url=page_url)
            if parsed:
                log.info(
                    "Parsed report: fio=%r phone=%r email=%r inn=%r",
                    parsed.fio,
                    parsed.phone,
                    parsed.email,
                    parsed.inn,
                )
                return parsed, None, None, None

            body_path, html_path, fixed_body_path = save_report_debug_artifacts(
                debug_dir,
                requested_inn=requested_inn,
                page_title=fixed_page_title,
                page_url=page_url,
                body_text=body_text,
                fixed_body_text=fixed_body_text,
                html_text=html_text,
            )
            log.warning(
                "Report parse failed. Debug artifacts saved: body=%s html=%s fixed_body=%s",
                body_path,
                html_path,
                fixed_body_path,
            )
            return None, body_path, html_path, fixed_body_path
        finally:
            await browser.close()


def build_result_row(state: QueryState) -> dict[str, str | None]:
    return {
        "requested_inn": state.requested_inn,
        "result_status": state.result_status,
        "status_message": state.status_message,
        "source_company_inn": None,
        "source_company_name": None,
        "last_company_inn": None,
        "last_company_name": None,
        "director_name": None,
        "director_inn": None,
        "person_fio": state.person.fio if state.person else None,
        "phone": state.person.phone if state.person else None,
        "email": state.person.email if state.person else None,
        "person_inn": state.person.inn if state.person else None,
    }


def append_result_csv(results_csv: Path, row: dict[str, str | None]) -> None:
    results_csv.parent.mkdir(parents=True, exist_ok=True)
    write_header = not results_csv.exists()
    with results_csv.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=RESULT_FIELDNAMES)
        if write_header:
            writer.writeheader()
        writer.writerow(row)


def append_result_xlsx(results_xlsx: Path, row: dict[str, str | None]) -> None:
    results_xlsx.parent.mkdir(parents=True, exist_ok=True)
    if results_xlsx.exists():
        workbook = load_workbook(results_xlsx)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "results"
        worksheet.append(RESULT_FIELDNAMES)

    worksheet.append([row.get(field) for field in RESULT_FIELDNAMES])
    workbook.save(results_xlsx)


def append_result(results_csv: Path, results_xlsx: Path, state: QueryState) -> None:
    row = build_result_row(state)
    append_result_csv(results_csv, row)
    append_result_xlsx(results_xlsx, row)


async def resolve_query(
    state: QueryState,
    log: logging.Logger,
    *,
    headless: bool,
    debug_dir: Path,
) -> bool:
    kind, payload = await wait_for_report_message(state, log)
    if kind is None:
        set_failure(
            state,
            status="no_response",
            message="После /inn бот не вернул ссылку на web-отчёт",
        )
        return False

    if kind == "not_found":
        not_found_message = payload
        assert isinstance(not_found_message, str)
        set_failure(
            state,
            status="not_found",
            message=not_found_message,
            error="По этому ИНН бот ничего не нашёл",
        )
        return False

    message = payload
    report_button = extract_report_button(message)
    if report_button is None or not report_button.url:
        set_failure(
            state,
            status="report_link_missing",
            message="Бот прислал сообщение про отчёт, но ссылка в кнопке не найдена",
        )
        return False

    print(f"[report] opening link: {report_button.url}")
    try:
        person, body_path, html_path, fixed_body_path = await fetch_person_from_report(
            report_button.url,
            requested_inn=state.requested_inn,
            headless=headless,
            debug_dir=debug_dir,
            log=log,
        )
    except PlaywrightTimeoutError:
        set_failure(
            state,
            status="web_open_failed",
            message="Не удалось дождаться загрузки web-отчёта",
        )
        return False
    except Exception as exc:
        set_failure(
            state,
            status="web_open_failed",
            message=f"Ошибка при открытии web-отчёта: {exc}",
        )
        return False

    if not person:
        set_failure(
            state,
            status="page_parse_failed",
            message=(
                "Страница отчёта открылась, но данные распарсить не удалось. "
                f"Debug body: {body_path}. Debug html: {html_path}. "
                f"Debug fixed body: {fixed_body_path}"
            ),
        )
        return False

    if not person.phone:
        state.person = person
        set_failure(
            state,
            status="phone_not_found",
            message="Web-отчёт открыт, но телефон на странице не найден",
        )
        return False

    state.person = person
    state.result_status = "found"
    state.status_message = "Телефон получен из web-отчёта"
    return True


async def run_single_query(
    client: TelegramClient,
    bot_entity,
    inn: str,
    *,
    log: logging.Logger | None = None,
    results_csv: Path | None = None,
    results_xlsx: Path | None = None,
    persist: bool = False,
    echo: bool = True,
    timeout_seconds: int = QUERY_TIMEOUT_SECONDS + 90,
    headless: bool = True,
    debug_dir: Path | None = None,
) -> QueryState:
    query_log = log or logging.getLogger("ip_phone")
    report_debug_dir = debug_dir or Path("report_debug")
    state = QueryState(requested_inn=inn)

    async def handle_bot_message(event, prefix: str) -> None:
        message = event.message
        if echo:
            print_incoming(prefix, message)
        state.queue.put_nowait(message)

    async def on_new_message(event):
        await handle_bot_message(event, "<")

    async def on_edited_message(event):
        await handle_bot_message(event, "< [edit]")

    new_message_builder = events.NewMessage(from_users=bot_entity)
    edited_message_builder = events.MessageEdited(from_users=bot_entity)
    client.add_event_handler(on_new_message, new_message_builder)
    client.add_event_handler(on_edited_message, edited_message_builder)

    try:
        command = f"/inn {inn}"
        if echo:
            print(f"[you] {command}")
        await client.send_message(bot_entity, command)

        try:
            found = await asyncio.wait_for(
                resolve_query(state, query_log, headless=headless, debug_dir=report_debug_dir),
                timeout=timeout_seconds,
            )
        except asyncio.TimeoutError:
            set_failure(
                state,
                status="timeout",
                message="Превышено время ожидания результата по ИП",
            )
        else:
            if found and state.result_status == "pending":
                state.result_status = "found"

        if persist:
            if results_csv is None or results_xlsx is None:
                raise RuntimeError("results_csv/results_xlsx are required when persist=True")
            append_result(results_csv, results_xlsx, state)

        return state
    finally:
        client.remove_event_handler(on_new_message, new_message_builder)
        client.remove_event_handler(on_edited_message, edited_message_builder)


async def main() -> None:
    log = setup_logging()
    api_id, api_hash, session_name, bot_username, results_csv, results_xlsx, headless, debug_dir, bot_message_echo = load_config()
    client = build_telegram_client(session_name, api_id, api_hash)
    current_query: QueryState | None = None

    try:
        await client.connect()
        if not await client.is_user_authorized():
            raise RuntimeError(
                "Telegram session is not authorized. Run `python qr_login.py` first to create "
                f"`{session_name}.session`."
            )

        bot_entity = await client.get_entity(bot_username)
        log.info("Connected to bot %s", bot_username)

        async def handle_bot_message(event, prefix: str) -> None:
            nonlocal current_query
            message = event.message
            if bot_message_echo:
                print_incoming(prefix, message)
            if current_query is not None:
                current_query.queue.put_nowait(message)

        @client.on(events.NewMessage(from_users=bot_entity))
        async def on_new_message(event):
            await handle_bot_message(event, "<")

        @client.on(events.MessageEdited(from_users=bot_entity))
        async def on_edited_message(event):
            await handle_bot_message(event, "< [edit]")

        print("Ready. Enter INN of IP or /exit.\n")

        while True:
            user_input = (await ainput("> ")).strip()
            if not user_input:
                continue
            if user_input == "/exit":
                break

            inn = normalize_inn(user_input)
            if not inn:
                print("[warn] Could not parse INN from input")
                continue

            current_query = QueryState(requested_inn=inn)
            command = f"/inn {inn}"
            if bot_message_echo:
                print(f"[you] {command}")
            await client.send_message(bot_entity, command)

            try:
                found = await asyncio.wait_for(
                    resolve_query(current_query, log, headless=headless, debug_dir=debug_dir),
                    timeout=QUERY_TIMEOUT_SECONDS + 90,
                )
            except asyncio.TimeoutError:
                set_failure(
                    current_query,
                    status="timeout",
                    message="Превышено время ожидания результата по ИП",
                )
                append_result(results_csv, results_xlsx, current_query)
                print("\n[result]")
                print(f"requested_inn: {current_query.requested_inn}")
                print(f"status: {current_query.result_status}")
                print(f"message: {current_query.status_message}")
                print(f"saved_to: {results_csv}")
                print(f"saved_to: {results_xlsx}")
                print()
                current_query = None
                continue

            append_result(results_csv, results_xlsx, current_query)

            print("\n[result]")
            print(f"requested_inn: {current_query.requested_inn}")
            print(f"status: {current_query.result_status}")
            print(f"message: {current_query.status_message or current_query.error or ''}")
            print(f"person: {current_query.person.fio if current_query.person else 'unknown'}")
            print(f"phone: {current_query.person.phone if current_query.person else 'not found'}")
            print(f"email: {current_query.person.email if current_query.person else 'not found'}")
            print(f"saved_to: {results_csv}")
            print(f"saved_to: {results_xlsx}")
            print()

            current_query = None
    finally:
        if client.is_connected():
            await client.disconnect()


if __name__ == "__main__":
    asyncio.run(main())
