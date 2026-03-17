import asyncio
import csv
import logging
import os
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from telethon import TelegramClient, events, helpers
from telethon.tl import types

load_dotenv()

NOT_FOUND_RE = re.compile(r"к сожалению,\s*по данному запросу ничего не найдено", re.IGNORECASE)

RESULT_FIELDNAMES = [
    "requested_phone",
    "result_status",
    "status_message",
    "phone",
    "operator",
    "region",
    "country",
    "fio",
    "birth_date",
    "age",
    "telegram",
    "email",
    "inn",
    "vk_text",
    "vk_urls",
    "instagram_text",
    "instagram_urls",
    "ok_text",
    "ok_urls",
]

QUERY_TIMEOUT_SECONDS = 90


@dataclass
class PhoneSummary:
    phone: str | None = None
    operator: str | None = None
    region: str | None = None
    country: str | None = None
    fio: str | None = None
    birth_date: str | None = None
    age: str | None = None
    telegram: str | None = None
    email: str | None = None
    inn: str | None = None
    vk_text: str | None = None
    vk_urls: str | None = None
    instagram_text: str | None = None
    instagram_urls: str | None = None
    ok_text: str | None = None
    ok_urls: str | None = None
    raw_text: str | None = None


@dataclass
class QueryState:
    requested_phone: str
    queue: asyncio.Queue[Any] = field(default_factory=asyncio.Queue)
    result_status: str = "pending"
    status_message: str | None = None
    summary: PhoneSummary | None = None
    error: str | None = None


def setup_logging() -> logging.Logger:
    level_name = os.getenv("LOG_LEVEL", "INFO").strip().upper()
    level = getattr(logging, level_name, logging.INFO)
    logging.basicConfig(
        level=level,
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    )
    logging.getLogger("telethon").setLevel(logging.WARNING)
    return logging.getLogger("phone_summary")


def get_required_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise RuntimeError(f"Missing required .env value: {name}")
    return value


def load_config() -> tuple[int, str, str, str, Path, Path]:
    api_id = int(get_required_env("API_ID"))
    api_hash = get_required_env("API_HASH")
    session_name = get_required_env("SESSION_NAME")
    bot_username = get_required_env("BOT")
    results_csv = Path(
        os.getenv("PHONE_SUMMARY_RESULTS_CSV", "phone_summary_results.csv").strip()
        or "phone_summary_results.csv"
    )
    default_xlsx = results_csv.with_suffix(".xlsx")
    results_xlsx = Path(
        os.getenv("PHONE_SUMMARY_RESULTS_XLSX", str(default_xlsx)).strip()
        or str(default_xlsx)
    )
    return api_id, api_hash, session_name, bot_username, results_csv, results_xlsx


def set_failure(state: QueryState, *, status: str, message: str, error: str | None = None) -> None:
    if state.result_status == "found":
        return
    state.result_status = status
    state.status_message = message
    state.error = error or message


async def ainput(prompt: str) -> str:
    return await asyncio.to_thread(input, prompt)


def normalize_phone_input(user_input: str) -> str | None:
    digits = re.sub(r"\D", "", user_input)
    if len(digits) == 10:
        return "7" + digits
    if 11 <= len(digits) <= 15:
        return digits
    return None


def get_message_text(message) -> str:
    return (getattr(message, "raw_text", None) or getattr(message, "text", None) or "").strip()


def print_incoming(prefix: str, message) -> None:
    text = get_message_text(message) or "[empty message]"
    print(f"\n{prefix} {text}")
    print("> ", end="", flush=True)


def parse_not_found_message(text: str) -> str | None:
    if NOT_FOUND_RE.search(text):
        return text.strip()
    return None


def normalize_phone_value(raw_phone: str | None) -> str | None:
    if not raw_phone:
        return None
    digits = re.sub(r"\D", "", raw_phone)
    if digits.startswith("7") and len(digits) > 11:
        digits = digits[:11]
    elif digits.startswith("8") and len(digits) > 11:
        digits = digits[:11]
    return digits or None


def is_label_line(line: str, labels: list[str]) -> bool:
    stripped = line.strip()
    for label in labels:
        if re.match(rf"^[^A-Za-zА-Яа-я0-9]*{re.escape(label)}\s*:", stripped, flags=re.IGNORECASE):
            return True
    return False


def extract_labeled_fields(text: str) -> dict[str, str]:
    labels = [
        "Телефон",
        "Оператор",
        "Регион",
        "Страна",
        "ФИО",
        "Дата рождения",
        "Возраст",
        "Telegram",
        "E-mail",
        "Email",
        "ИНН",
    ]
    multi_line_labels = {"Telegram", "E-mail", "Email"}

    result: dict[str, str] = {}
    current_label: str | None = None

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            current_label = None
            continue

        matched_label: str | None = None
        matched_value: str | None = None
        for label in labels:
            match = re.match(
                rf"^[^A-Za-zА-Яа-я0-9]*{re.escape(label)}\s*:\s*(.*)$",
                line,
                flags=re.IGNORECASE,
            )
            if match:
                matched_label = label
                matched_value = match.group(1).strip()
                break

        if matched_label:
            current_label = matched_label
            canonical_label = "Email" if matched_label in {"E-mail", "Email"} else matched_label
            result[canonical_label] = matched_value or ""
            continue

        if current_label in multi_line_labels and not is_label_line(line, labels):
            canonical_label = "Email" if current_label in {"E-mail", "Email"} else current_label
            existing = result.get(canonical_label, "")
            result[canonical_label] = (existing + " " + line).strip()

    return result


def get_entity_text(text: str, entity) -> str:
    surrogate_text = helpers.add_surrogate(text)
    entity_surrogate_text = surrogate_text[entity.offset : entity.offset + entity.length]
    return helpers.del_surrogate(entity_surrogate_text)


def join_unique(items: list[str]) -> str | None:
    cleaned: list[str] = []
    seen: set[str] = set()
    for item in items:
        value = item.strip()
        if not value or value in seen:
            continue
        seen.add(value)
        cleaned.append(value)
    return " | ".join(cleaned) if cleaned else None


def extract_social_links(message, text: str) -> dict[str, list[str]]:
    result = {
        "vk_text": [],
        "vk_urls": [],
        "instagram_text": [],
        "instagram_urls": [],
        "ok_text": [],
        "ok_urls": [],
    }

    entities = getattr(message, "entities", None) or []
    lines = text.splitlines()

    for entity in entities:
        entity_text = get_entity_text(text, entity).strip()
        if not entity_text:
            continue

        line = ""
        for raw_line in lines:
            if entity_text in raw_line:
                line = raw_line.strip()
                break

        lower_line = line.casefold()
        url: str | None = None

        if isinstance(entity, types.MessageEntityTextUrl):
            url = entity.url
        elif isinstance(entity, types.MessageEntityUrl):
            url = entity_text

        if "вконтакте" in lower_line or (url and "vk.com" in url.casefold()):
            result["vk_text"].append(entity_text)
            if url:
                result["vk_urls"].append(url)
        elif "instagram" in lower_line or (url and "instagram.com" in url.casefold()):
            result["instagram_text"].append(entity_text)
            if url:
                result["instagram_urls"].append(url)
        elif "одноклассники" in lower_line or (url and "ok.ru" in url.casefold()):
            result["ok_text"].append(entity_text)
            if url:
                result["ok_urls"].append(url)

    return result


def parse_phone_summary(message) -> PhoneSummary | None:
    text = get_message_text(message)
    fields = extract_labeled_fields(text)
    if "Телефон" not in fields and "ФИО" not in fields:
        return None

    birth_date = fields.get("Дата рождения")
    if birth_date:
        birth_date = birth_date.split("(")[0].strip()

    age = fields.get("Возраст")
    if age:
        age = re.sub(r"\D", "", age) or age

    social = extract_social_links(message, text)

    summary = PhoneSummary(
        phone=normalize_phone_value(fields.get("Телефон")),
        operator=fields.get("Оператор") or None,
        region=fields.get("Регион") or None,
        country=fields.get("Страна") or None,
        fio=fields.get("ФИО") or None,
        birth_date=birth_date or None,
        age=age or None,
        telegram=fields.get("Telegram") or None,
        email=fields.get("Email") or None,
        inn=fields.get("ИНН") or None,
        vk_text=join_unique(social["vk_text"]),
        vk_urls=join_unique(social["vk_urls"]),
        instagram_text=join_unique(social["instagram_text"]),
        instagram_urls=join_unique(social["instagram_urls"]),
        ok_text=join_unique(social["ok_text"]),
        ok_urls=join_unique(social["ok_urls"]),
        raw_text=text,
    )
    return summary


async def wait_for_summary_message(state: QueryState, log: logging.Logger) -> tuple[str | None, Any | None]:
    loop = asyncio.get_running_loop()
    deadline = loop.time() + QUERY_TIMEOUT_SECONDS

    while True:
        remaining = deadline - loop.time()
        if remaining <= 0:
            return None, None

        try:
            message = await asyncio.wait_for(state.queue.get(), timeout=remaining)
        except asyncio.TimeoutError:
            return None, None

        text = get_message_text(message)
        not_found_message = parse_not_found_message(text)
        if not_found_message:
            log.info("Received not-found response for current query")
            return "not_found", not_found_message

        summary = parse_phone_summary(message)
        if summary:
            log.info("Received phone summary: fio=%r phone=%r", summary.fio, summary.phone)
            return "summary", summary


def build_result_row(state: QueryState) -> dict[str, str | None]:
    summary = state.summary
    return {
        "requested_phone": state.requested_phone,
        "result_status": state.result_status,
        "status_message": state.status_message,
        "phone": summary.phone if summary else None,
        "operator": summary.operator if summary else None,
        "region": summary.region if summary else None,
        "country": summary.country if summary else None,
        "fio": summary.fio if summary else None,
        "birth_date": summary.birth_date if summary else None,
        "age": summary.age if summary else None,
        "telegram": summary.telegram if summary else None,
        "email": summary.email if summary else None,
        "inn": summary.inn if summary else None,
        "vk_text": summary.vk_text if summary else None,
        "vk_urls": summary.vk_urls if summary else None,
        "instagram_text": summary.instagram_text if summary else None,
        "instagram_urls": summary.instagram_urls if summary else None,
        "ok_text": summary.ok_text if summary else None,
        "ok_urls": summary.ok_urls if summary else None,
    }


def append_result_csv(results_csv: Path, row: dict[str, str | None]) -> None:
    results_csv.parent.mkdir(parents=True, exist_ok=True)
    write_header = not results_csv.exists()
    with results_csv.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=RESULT_FIELDNAMES)
        if write_header:
            writer.writeheader()
        writer.writerow(row)


def append_result_xlsx(results_xlsx: Path, row: dict[str, str | None]) -> None:
    results_xlsx.parent.mkdir(parents=True, exist_ok=True)
    if results_xlsx.exists():
        workbook = load_workbook(results_xlsx)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "phone_summary"
        worksheet.append(RESULT_FIELDNAMES)

    worksheet.append([row.get(field) for field in RESULT_FIELDNAMES])
    workbook.save(results_xlsx)


def append_result(results_csv: Path, results_xlsx: Path, state: QueryState) -> None:
    row = build_result_row(state)
    append_result_csv(results_csv, row)
    append_result_xlsx(results_xlsx, row)


async def resolve_query(state: QueryState, log: logging.Logger) -> bool:
    kind, payload = await wait_for_summary_message(state, log)
    if kind is None:
        set_failure(
            state,
            status="no_response",
            message="После отправки телефона бот не вернул полезный ответ",
        )
        return False

    if kind == "not_found":
        not_found_message = payload
        assert isinstance(not_found_message, str)
        set_failure(
            state,
            status="not_found",
            message=not_found_message,
            error="По этому телефону бот ничего не нашёл",
        )
        return False

    summary = payload
    assert isinstance(summary, PhoneSummary)
    state.summary = summary
    state.result_status = "found"
    state.status_message = "Краткая сводка получена"
    return True


async def run_single_query(
    client: TelegramClient,
    bot_entity,
    phone: str,
    *,
    log: logging.Logger | None = None,
    results_csv: Path | None = None,
    results_xlsx: Path | None = None,
    persist: bool = False,
    echo: bool = True,
    timeout_seconds: int = QUERY_TIMEOUT_SECONDS + 30,
) -> QueryState:
    query_log = log or logging.getLogger("phone_summary")
    state = QueryState(requested_phone=phone)

    async def handle_bot_message(event, prefix: str) -> None:
        message = event.message
        if echo:
            print_incoming(prefix, message)
        state.queue.put_nowait(message)

    async def on_new_message(event):
        await handle_bot_message(event, "<")

    async def on_edited_message(event):
        await handle_bot_message(event, "< [edit]")

    new_message_builder = events.NewMessage(from_users=bot_entity)
    edited_message_builder = events.MessageEdited(from_users=bot_entity)
    client.add_event_handler(on_new_message, new_message_builder)
    client.add_event_handler(on_edited_message, edited_message_builder)

    try:
        if echo:
            print(f"[you] {phone}")
        await client.send_message(bot_entity, phone)

        try:
            found = await asyncio.wait_for(resolve_query(state, query_log), timeout=timeout_seconds)
        except asyncio.TimeoutError:
            set_failure(
                state,
                status="timeout",
                message="Превышено время ожидания результата по телефону",
            )
        else:
            if found and state.result_status == "pending":
                state.result_status = "found"

        if persist:
            if results_csv is None or results_xlsx is None:
                raise RuntimeError("results_csv/results_xlsx are required when persist=True")
            append_result(results_csv, results_xlsx, state)

        return state
    finally:
        client.remove_event_handler(on_new_message, new_message_builder)
        client.remove_event_handler(on_edited_message, edited_message_builder)


async def main() -> None:
    log = setup_logging()
    api_id, api_hash, session_name, bot_username, results_csv, results_xlsx = load_config()
    client = TelegramClient(session_name, api_id, api_hash)
    current_query: QueryState | None = None

    try:
        await client.connect()
        if not await client.is_user_authorized():
            raise RuntimeError(
                "Telegram session is not authorized. Run `python qr_login.py` first to create "
                f"`{session_name}.session`."
            )

        bot_entity = await client.get_entity(bot_username)
        log.info("Connected to bot %s", bot_username)

        async def handle_bot_message(event, prefix: str) -> None:
            nonlocal current_query
            message = event.message
            print_incoming(prefix, message)
            if current_query is not None:
                current_query.queue.put_nowait(message)

        @client.on(events.NewMessage(from_users=bot_entity))
        async def on_new_message(event):
            await handle_bot_message(event, "<")

        @client.on(events.MessageEdited(from_users=bot_entity))
        async def on_edited_message(event):
            await handle_bot_message(event, "< [edit]")

        print("Ready. Enter phone or /exit.\n")

        while True:
            user_input = (await ainput("> ")).strip()
            if not user_input:
                continue
            if user_input == "/exit":
                break

            phone = normalize_phone_input(user_input)
            if not phone:
                print("[warn] Could not parse phone from input")
                continue

            current_query = QueryState(requested_phone=phone)
            print(f"[you] {phone}")
            await client.send_message(bot_entity, phone)

            try:
                found = await asyncio.wait_for(resolve_query(current_query, log), timeout=QUERY_TIMEOUT_SECONDS + 30)
            except asyncio.TimeoutError:
                set_failure(
                    current_query,
                    status="timeout",
                    message="Превышено время ожидания результата по телефону",
                )
                append_result(results_csv, results_xlsx, current_query)
                print("\n[result]")
                print(f"requested_phone: {current_query.requested_phone}")
                print(f"status: {current_query.result_status}")
                print(f"message: {current_query.status_message}")
                print(f"saved_to: {results_csv}")
                print(f"saved_to: {results_xlsx}")
                print()
                current_query = None
                continue

            append_result(results_csv, results_xlsx, current_query)

            print("\n[result]")
            print(f"requested_phone: {current_query.requested_phone}")
            print(f"status: {current_query.result_status}")
            print(f"message: {current_query.status_message or current_query.error or ''}")
            print(f"fio: {current_query.summary.fio if current_query.summary else 'unknown'}")
            print(f"phone: {current_query.summary.phone if current_query.summary else 'not found'}")
            print(f"email: {current_query.summary.email if current_query.summary else 'not found'}")
            print(f"inn: {current_query.summary.inn if current_query.summary else 'not found'}")
            print(f"vk_urls: {current_query.summary.vk_urls if current_query.summary else 'not found'}")
            print(f"instagram_urls: {current_query.summary.instagram_urls if current_query.summary else 'not found'}")
            print(f"ok_urls: {current_query.summary.ok_urls if current_query.summary else 'not found'}")
            print(f"saved_to: {results_csv}")
            print(f"saved_to: {results_xlsx}")
            print()

            current_query = None
    finally:
        if client.is_connected():
            await client.disconnect()


if __name__ == "__main__":
    asyncio.run(main())
